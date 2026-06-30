import csv
import io
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def export_bibtex(papers: List[Dict[str, Any]]) -> str:
    lines = []
    for p in papers:
        key = (p.get("doi") or p.get("title", "unknown")).replace("/", "_").replace(" ", "")[:40]
        entry = [f"@article{{{key},"]
        if p.get("title"):
            entry.append(f'  title = {{{p["title"]}}},')
        if p.get("authors"):
            entry.append(f'  author = {{{p["authors"]}}},')
        if p.get("year"):
            entry.append(f'  year = {{{p["year"]}}},')
        if p.get("journal"):
            entry.append(f'  journal = {{{p["journal"]}}},')
        if p.get("doi"):
            entry.append(f'  doi = {{{p["doi"]}}},')
        entry.append("}")
        lines.append("\n".join(entry))
    return "\n\n".join(lines)


def export_ris(papers: List[Dict[str, Any]]) -> str:
    lines = []
    for p in papers:
        entry = ["TY  - JOUR"]
        if p.get("title"):
            entry.append(f'TI  - {p["title"]}')
        if p.get("authors"):
            for author in p["authors"].split(";"):
                a = author.strip()
                if a:
                    entry.append(f"AU  - {a}")
        if p.get("year"):
            entry.append(f'PY  - {p["year"]}')
        if p.get("journal"):
            entry.append(f'JO  - {p["journal"]}')
        if p.get("doi"):
            entry.append(f'DO  - {p["doi"]}')
        if p.get("abstract"):
            entry.append(f'AB  - {p["abstract"]}')
        entry.append("ER  -")
        lines.append("\n".join(entry))
    return "\n\n".join(lines)


def export_csv(papers: List[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=["title", "authors", "year", "journal", "doi", "citation_count", "is_open_access", "topics"],
        extrasaction="ignore",
    )
    writer.writeheader()
    for p in papers:
        writer.writerow(p)
    return buf.getvalue().encode("utf-8")


def export_excel(papers: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature Matrix"

    headers = ["Title", "Authors", "Year", "Journal", "DOI", "Citations", "Open Access", "Topics"]
    fields = ["title", "authors", "year", "journal", "doi", "citation_count", "is_open_access", "topics"]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, paper in enumerate(papers, start=2):
        for col_idx, field in enumerate(fields, start=1):
            val = paper.get(field)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [60, 40, 8, 35, 40, 12, 14, 50]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
